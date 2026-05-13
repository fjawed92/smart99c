"""Excel bulk import for products and color variants.

Template columns (one row = one variant of a product, or one row per product
when ``color_name`` is blank):

    product_slug, product_name, category_name, short_description, description,
    price, compare_price, cost_price, product_sku, weight,
    is_active, is_featured, track_inventory, stock_quantity,
    color_name, color_hex, variant_sku, variant_price_override,
    variant_cost_price, variant_stock_quantity

Blank cells are treated as "leave unchanged" for existing records (not
"clear"). Boolean cells accept TRUE/FALSE, 1/0, yes/no, on/off.
"""
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

from app.extensions import db
from app.models import Category, Product, ProductVariant
from app.helpers import generate_slug


COLUMNS = [
    'product_slug', 'product_name', 'category_name', 'short_description', 'description',
    'price', 'compare_price', 'cost_price', 'product_sku', 'weight',
    'is_active', 'is_featured', 'track_inventory', 'stock_quantity',
    'color_name', 'color_hex', 'variant_sku', 'variant_price_override',
    'variant_cost_price', 'variant_stock_quantity',
]


def _is_blank(value):
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _as_str(value):
    if _is_blank(value):
        return None
    return str(value).strip()


def _as_decimal(value):
    if _is_blank(value):
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f'invalid number: {value!r}')


def _as_int(value):
    if _is_blank(value):
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        raise ValueError(f'invalid integer: {value!r}')


_TRUE = {'1', 'true', 'yes', 'y', 'on', 't'}
_FALSE = {'0', 'false', 'no', 'n', 'off', 'f'}


def _as_bool(value):
    if _is_blank(value):
        return None
    s = str(value).strip().lower()
    if s in _TRUE:
        return True
    if s in _FALSE:
        return False
    raise ValueError(f'invalid boolean: {value!r}')


def _unique_slug(base, existing_id=None):
    slug = base
    suffix = 0
    while True:
        candidate = slug if suffix == 0 else f'{slug}-{suffix}'
        clash = Product.query.filter_by(slug=candidate).first()
        if not clash or (existing_id and clash.id == existing_id):
            return candidate
        suffix += 1


def _resolve_category(name, cache):
    if not name:
        return None, None
    key = name.strip().lower()
    if key in cache:
        return cache[key], None
    cat = Category.query.filter(db.func.lower(Category.name) == key).first()
    cache[key] = cat
    if not cat:
        return None, f'category "{name}" not found'
    return cat, None


def _find_product(slug, name):
    if slug:
        p = Product.query.filter_by(slug=slug).first()
        if p:
            return p
    if name:
        return Product.query.filter(db.func.lower(Product.name) == name.strip().lower()).first()
    return None


def _find_variant(product, variant_sku, color_name):
    if variant_sku:
        v = ProductVariant.query.filter_by(product_id=product.id, sku=variant_sku).first()
        if v:
            return v
    if color_name:
        return ProductVariant.query.filter(
            ProductVariant.product_id == product.id,
            db.func.lower(ProductVariant.color_name) == color_name.strip().lower()
        ).first()
    return None


def import_workbook(file_storage):
    """Process an uploaded .xlsx FileStorage. Returns a result dict."""
    result = {
        'created': 0,
        'updated': 0,
        'variants_created': 0,
        'variants_updated': 0,
        'skipped': 0,
        'errors': [],
    }

    try:
        wb = load_workbook(filename=BytesIO(file_storage.read()), read_only=True, data_only=True)
    except Exception as e:
        result['errors'].append((0, f'Could not read workbook: {e}'))
        return result

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip().lower() if c is not None else '' for c in next(rows)]
    except StopIteration:
        result['errors'].append((0, 'File is empty.'))
        return result

    missing = [c for c in COLUMNS if c not in header]
    if missing:
        result['errors'].append((1, f'Missing columns: {", ".join(missing)}'))
        return result

    idx = {name: header.index(name) for name in COLUMNS}
    cat_cache = {}

    def cell(row, name):
        i = idx[name]
        return row[i] if i < len(row) else None

    row_number = 1
    for row in rows:
        row_number += 1
        if row is None or all(_is_blank(c) for c in row):
            continue

        savepoint = db.session.begin_nested()
        try:
            slug_in = _as_str(cell(row, 'product_slug'))
            name_in = _as_str(cell(row, 'product_name'))
            color_in = _as_str(cell(row, 'color_name'))

            product = _find_product(slug_in, name_in)
            created = False
            if product is None:
                if not name_in:
                    result['skipped'] += 1
                    result['errors'].append((row_number, 'product_name is required to create a new product'))
                    continue
                price_in = _as_decimal(cell(row, 'price'))
                if price_in is None:
                    result['skipped'] += 1
                    result['errors'].append((row_number, 'price is required to create a new product'))
                    continue
                base_slug = slug_in or generate_slug(name_in)
                product = Product(
                    name=name_in,
                    slug=_unique_slug(base_slug),
                    price=price_in,
                    stock_quantity=0,
                    track_inventory=True,
                    is_active=True,
                )
                db.session.add(product)
                db.session.flush()
                created = True
                result['created'] += 1

            # Update product fields when supplied
            short_desc = _as_str(cell(row, 'short_description'))
            if short_desc is not None:
                product.short_description = short_desc
            desc = _as_str(cell(row, 'description'))
            if desc is not None:
                product.description = desc
            price = _as_decimal(cell(row, 'price'))
            if price is not None:
                product.price = price
            compare_price = _as_decimal(cell(row, 'compare_price'))
            if compare_price is not None:
                product.compare_price = compare_price
            cost_price = _as_decimal(cell(row, 'cost_price'))
            if cost_price is not None:
                product.cost_price = cost_price
            psku = _as_str(cell(row, 'product_sku'))
            if psku is not None:
                product.sku = psku
            weight = _as_decimal(cell(row, 'weight'))
            if weight is not None:
                product.weight = weight
            cat_name = _as_str(cell(row, 'category_name'))
            if cat_name is not None:
                cat, cat_err = _resolve_category(cat_name, cat_cache)
                if cat_err:
                    result['errors'].append((row_number, cat_err))
                else:
                    product.category_id = cat.id
            is_active = _as_bool(cell(row, 'is_active'))
            if is_active is not None:
                product.is_active = is_active
            is_featured = _as_bool(cell(row, 'is_featured'))
            if is_featured is not None:
                product.is_featured = is_featured
            track_inventory = _as_bool(cell(row, 'track_inventory'))
            if track_inventory is not None:
                product.track_inventory = track_inventory

            if color_in:
                v_sku = _as_str(cell(row, 'variant_sku'))
                variant = _find_variant(product, v_sku, color_in)
                v_created = False
                if variant is None:
                    variant = ProductVariant(
                        product_id=product.id,
                        color_name=color_in,
                        stock_quantity=0,
                    )
                    db.session.add(variant)
                    db.session.flush()
                    v_created = True
                else:
                    variant.color_name = color_in

                color_hex = _as_str(cell(row, 'color_hex'))
                if color_hex is not None:
                    variant.color_hex = color_hex
                if v_sku is not None:
                    variant.sku = v_sku
                v_price = _as_decimal(cell(row, 'variant_price_override'))
                if v_price is not None:
                    variant.price_override = v_price
                v_cost = _as_decimal(cell(row, 'variant_cost_price'))
                if v_cost is not None:
                    variant.cost_price = v_cost
                v_stock = _as_int(cell(row, 'variant_stock_quantity'))
                if v_stock is not None:
                    variant.stock_quantity = max(0, v_stock)

                if v_created:
                    result['variants_created'] += 1
                else:
                    result['variants_updated'] += 1
            else:
                stock = _as_int(cell(row, 'stock_quantity'))
                if stock is not None:
                    product.stock_quantity = max(0, stock)

            if not created:
                result['updated'] += 1
            savepoint.commit()

        except ValueError as e:
            savepoint.rollback()
            result['skipped'] += 1
            result['errors'].append((row_number, str(e)))
            continue
        except Exception as e:
            savepoint.rollback()
            result['skipped'] += 1
            result['errors'].append((row_number, f'unexpected error: {e}'))
            continue

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        result['errors'].append((0, f'Commit failed: {e}'))

    return result
