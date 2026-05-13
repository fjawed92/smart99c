"""Excel export for products and color variants.

Generates an .xlsx workbook matching the import template format. The same
file can be edited and re-uploaded via the import flow.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models import Category, Product
from app.services.product_import import COLUMNS


def _decimal_to_value(d):
    if d is None:
        return None
    return float(d)


def _bool_to_value(b):
    if b is None:
        return None
    return 'TRUE' if b else 'FALSE'


def _row_for_product(product, variant=None):
    return [
        product.slug,
        product.name,
        product.category.name if product.category else None,
        product.short_description,
        product.description,
        _decimal_to_value(product.price),
        _decimal_to_value(product.compare_price),
        _decimal_to_value(product.cost_price),
        product.sku,
        _decimal_to_value(product.weight),
        _bool_to_value(product.is_active),
        _bool_to_value(product.is_featured),
        _bool_to_value(product.track_inventory),
        product.stock_quantity if variant is None else None,
        variant.color_name if variant else None,
        variant.color_hex if variant else None,
        variant.sku if variant else None,
        _decimal_to_value(variant.price_override) if variant else None,
        _decimal_to_value(variant.cost_price) if variant else None,
        variant.stock_quantity if variant else None,
    ]


def _style_header(ws):
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='E8334A')
    align = Alignment(horizontal='left', vertical='center')
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align
    # Reasonable widths
    widths = {
        'product_slug': 22, 'product_name': 28, 'category_name': 18,
        'short_description': 30, 'description': 40, 'price': 9,
        'compare_price': 13, 'cost_price': 11, 'product_sku': 16,
        'weight': 8, 'is_active': 10, 'is_featured': 11,
        'track_inventory': 14, 'stock_quantity': 13,
        'color_name': 14, 'color_hex': 11, 'variant_sku': 16,
        'variant_price_override': 18, 'variant_cost_price': 16,
        'variant_stock_quantity': 18,
    }
    for col_idx, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = widths.get(name, 14)


def _append_category_sheet(wb):
    ws = wb.create_sheet('Categories')
    ws.cell(row=1, column=1, value='category_name').font = Font(bold=True)
    for i, cat in enumerate(Category.query.order_by(Category.name).all(), start=2):
        ws.cell(row=i, column=1, value=cat.name)
    ws.column_dimensions['A'].width = 30


def build_template_workbook():
    """Empty template with header + 2 example rows + category reference sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'
    _style_header(ws)

    example_rows = [
        # A product without variants
        ['', 'Example T-Shirt', 'Apparel', 'Soft cotton tee',
         'A comfy basic tee.', 9.99, None, 3.50, 'TEE-BASIC', 0.3,
         'TRUE', 'FALSE', 'TRUE', 50,
         '', '', '', '', '', ''],
        # A product with two color variants — same product_slug repeated
        ['example-mug', 'Example Mug', 'Kitchen', '11oz ceramic mug',
         'Holds 11oz of hot or cold beverage.', 6.99, 8.99, 2.10, 'MUG-11', 0.6,
         'TRUE', 'TRUE', 'TRUE', None,
         'Red', '#FF0000', 'MUG-11-RED', None, None, 25],
        ['example-mug', 'Example Mug', 'Kitchen', '', '',
         None, None, None, '', None,
         '', '', '', None,
         'Blue', '#0066FF', 'MUG-11-BLU', None, None, 30],
    ]
    for r in example_rows:
        ws.append(r)

    _append_category_sheet(wb)
    return wb


def build_catalog_workbook():
    """Full export of every product + variant."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'
    _style_header(ws)

    products = (Product.query
                .order_by(Product.name)
                .all())
    for p in products:
        if p.variants:
            for v in p.variants:
                ws.append(_row_for_product(p, v))
        else:
            ws.append(_row_for_product(p, None))

    _append_category_sheet(wb)
    return wb


def workbook_to_bytes(wb):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
