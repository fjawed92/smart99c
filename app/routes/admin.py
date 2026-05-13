import cloudinary
import cloudinary.uploader
from datetime import datetime, timedelta
from decimal import Decimal
from functools import wraps
from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, jsonify, abort, current_app, send_file)
from flask_login import login_required, current_user
from flask_wtf import FlaskForm
from wtforms import (StringField, TextAreaField, DecimalField, IntegerField,
                     BooleanField, SelectField, FileField, PasswordField, validators)
from app.extensions import db
from app.models import (Product, ProductImage, ProductVariant, Category, Order,
                        OrderItem, User, ShippingRate, SiteSettings, PaymentLink)
from app.helpers import (generate_slug, get_site_setting, set_site_setting,
                         encrypt_secret)

admin_bp = Blueprint('admin', __name__)


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return login_required(decorated)


# ─── Dashboard ───────────────────────────────────────────────────────────────

@admin_bp.route('/')
@admin_required
def dashboard():
    today = datetime.utcnow().date()
    month_start = today.replace(day=1)

    total_orders = Order.query.count()
    revenue_today = db.session.query(db.func.sum(Order.total))\
        .filter(db.func.date(Order.created_at) == today,
                Order.status.notin_(['cancelled', 'refunded'])).scalar() or 0
    revenue_month = db.session.query(db.func.sum(Order.total))\
        .filter(Order.created_at >= month_start,
                Order.status.notin_(['cancelled', 'refunded'])).scalar() or 0
    total_products = Product.query.filter_by(is_active=True).count()
    low_stock = Product.query.filter(
        Product.track_inventory == True,
        Product.stock_quantity <= 5,
        Product.is_active == True
    ).count()
    new_customers = User.query.filter(
        db.func.date(User.created_at) >= month_start,
        User.is_admin == False
    ).count()
    recent_orders = Order.query.order_by(Order.created_at.desc()).limit(10).all()

    return render_template('admin/dashboard.html',
                           total_orders=total_orders,
                           revenue_today=revenue_today,
                           revenue_month=revenue_month,
                           total_products=total_products,
                           low_stock=low_stock,
                           new_customers=new_customers,
                           recent_orders=recent_orders)


# ─── Products ────────────────────────────────────────────────────────────────

class ProductForm(FlaskForm):
    name = StringField('Name', [validators.DataRequired(), validators.Length(max=255)])
    short_description = StringField('Short Description', [validators.Optional(), validators.Length(max=500)])
    description = TextAreaField('Description', [validators.Optional()])
    price = DecimalField('Price', [validators.DataRequired(), validators.NumberRange(min=0)], places=2)
    compare_price = DecimalField('Compare Price', [validators.Optional(), validators.NumberRange(min=0)], places=2)
    cost_price = DecimalField('Cost Price', [validators.Optional(), validators.NumberRange(min=0)], places=2)
    sku = StringField('SKU', [validators.Optional(), validators.Length(max=100)])
    stock_quantity = IntegerField('Stock', [validators.Optional(), validators.NumberRange(min=0)], default=0)
    track_inventory = BooleanField('Track Inventory', default=True)
    weight = DecimalField('Weight (lbs)', [validators.Optional(), validators.NumberRange(min=0)], places=2, default=0)
    category_id = SelectField('Category', coerce=int)
    is_active = BooleanField('Active', default=True)
    is_featured = BooleanField('Featured', default=False)


@admin_bp.route('/products')
@admin_required
def products():
    search = request.args.get('q', '')
    category_id = request.args.get('category_id', type=int)
    status = request.args.get('status', '')
    page = request.args.get('page', 1, type=int)

    query = Product.query
    if search:
        query = query.filter(Product.name.ilike(f'%{search}%'))
    if category_id:
        query = query.filter_by(category_id=category_id)
    if status == 'active':
        query = query.filter_by(is_active=True)
    elif status == 'inactive':
        query = query.filter_by(is_active=False)

    pagination = query.order_by(Product.created_at.desc()).paginate(page=page, per_page=20)
    categories = Category.query.order_by(Category.name).all()
    return render_template('admin/products.html',
                           pagination=pagination,
                           products=pagination.items,
                           categories=categories,
                           search=search,
                           category_id=category_id,
                           status=status)


@admin_bp.route('/products/new', methods=['GET', 'POST'])
@admin_required
def new_product():
    form = ProductForm()
    form.category_id.choices = [(0, '— No Category —')] + [
        (c.id, c.name) for c in Category.query.order_by(Category.name).all()
    ]
    if form.validate_on_submit():
        slug = generate_slug(form.name.data)
        existing = Product.query.filter_by(slug=slug).first()
        if existing:
            slug = f'{slug}-{int(datetime.utcnow().timestamp())}'

        product = Product(
            name=form.name.data,
            slug=slug,
            description=form.description.data,
            short_description=form.short_description.data,
            price=form.price.data,
            compare_price=form.compare_price.data or None,
            cost_price=form.cost_price.data or None,
            sku=form.sku.data or None,
            stock_quantity=form.stock_quantity.data or 0,
            track_inventory=form.track_inventory.data,
            weight=form.weight.data or 0,
            category_id=form.category_id.data or None,
            is_active=form.is_active.data,
            is_featured=form.is_featured.data,
        )
        db.session.add(product)
        db.session.flush()

        _handle_image_uploads(product, request.files.getlist('images'))

        db.session.commit()
        flash(f'Product "{product.name}" created!', 'success')
        return redirect(url_for('admin.products'))

    return render_template('admin/product_form.html', form=form, product=None)


@admin_bp.route('/products/<int:product_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_product(product_id):
    product = Product.query.get_or_404(product_id)
    form = ProductForm(obj=product)
    form.category_id.choices = [(0, '— No Category —')] + [
        (c.id, c.name) for c in Category.query.order_by(Category.name).all()
    ]

    if form.validate_on_submit():
        product.name = form.name.data
        product.description = form.description.data
        product.short_description = form.short_description.data
        product.price = form.price.data
        product.compare_price = form.compare_price.data or None
        product.cost_price = form.cost_price.data or None
        product.sku = form.sku.data or None
        product.stock_quantity = form.stock_quantity.data or 0
        product.track_inventory = form.track_inventory.data
        product.weight = form.weight.data or 0
        product.category_id = form.category_id.data or None
        product.is_active = form.is_active.data
        product.is_featured = form.is_featured.data

        _handle_image_uploads(product, request.files.getlist('images'))
        db.session.commit()
        flash(f'Product "{product.name}" updated!', 'success')
        return redirect(url_for('admin.products'))

    form.category_id.data = product.category_id or 0
    return render_template('admin/product_form.html', form=form, product=product)


@admin_bp.route('/products/<int:product_id>/delete', methods=['POST'])
@admin_required
def delete_product(product_id):
    product = Product.query.get_or_404(product_id)
    product.is_active = False
    db.session.commit()
    flash(f'Product "{product.name}" deactivated.', 'success')
    return redirect(url_for('admin.products'))


@admin_bp.route('/products/<int:product_id>/toggle', methods=['POST'])
@admin_required
def toggle_product(product_id):
    product = Product.query.get_or_404(product_id)
    product.is_active = not product.is_active
    db.session.commit()
    state = 'activated' if product.is_active else 'deactivated'
    return jsonify({'success': True, 'is_active': product.is_active,
                    'message': f'Product {state}.'})


@admin_bp.route('/products/bulk-action', methods=['POST'])
@admin_required
def bulk_product_action(action=None):
    action = request.form.get('action', '')
    ids = request.form.getlist('product_ids')
    try:
        ids = [int(i) for i in ids if str(i).isdigit()]
    except ValueError:
        ids = []
    if not ids:
        flash('Select at least one product.', 'error')
        return redirect(request.referrer or url_for('admin.products'))

    products = Product.query.filter(Product.id.in_(ids)).all()
    if action == 'activate':
        for p in products:
            p.is_active = True
        db.session.commit()
        flash(f'Activated {len(products)} product(s).', 'success')
    elif action == 'deactivate':
        for p in products:
            p.is_active = False
        db.session.commit()
        flash(f'Deactivated {len(products)} product(s).', 'success')
    elif action == 'feature':
        for p in products:
            p.is_featured = True
        db.session.commit()
        flash(f'Featured {len(products)} product(s).', 'success')
    elif action == 'unfeature':
        for p in products:
            p.is_featured = False
        db.session.commit()
        flash(f'Unfeatured {len(products)} product(s).', 'success')
    elif action == 'delete':
        # Soft delete: same as the per-row Delete button (deactivate).
        for p in products:
            p.is_active = False
        db.session.commit()
        flash(f'Deactivated {len(products)} product(s).', 'success')
    else:
        flash('Unknown action.', 'error')

    return redirect(request.referrer or url_for('admin.products'))


@admin_bp.route('/orders/export', methods=['GET'])
@admin_required
def orders_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from app.models import ShippingAddress

    wb = Workbook()
    ws = wb.active
    ws.title = 'Orders'

    header = ['Order #', 'Date', 'Status', 'Customer', 'Email', 'Phone',
              'Items', 'Subtotal', 'Tax', 'Shipping', 'Total', 'Payment Intent']
    for i, h in enumerate(header, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='E8334A')
        c.alignment = Alignment(horizontal='left', vertical='center')

    widths = [14, 18, 12, 24, 28, 16, 50, 10, 10, 11, 10, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    orders = Order.query.order_by(Order.created_at.desc()).all()
    for o in orders:
        items_str = '; '.join(
            f'{i.display_name} x{i.quantity} @ ${i.unit_price}' for i in o.items
        )
        addr = o.shipping_address
        ws.append([
            o.order_number,
            o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else '',
            o.status,
            addr.full_name if addr else (o.user.full_name if o.user else ''),
            addr.email if addr else (o.user.email if o.user else ''),
            addr.phone if addr else '',
            items_str,
            float(o.subtotal or 0),
            float(o.tax or 0),
            float(o.shipping_cost or 0),
            float(o.total or 0),
            o.stripe_payment_intent_id or '',
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.utcnow().strftime('%Y%m%d')
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'smart99c-orders-{stamp}.xlsx')


@admin_bp.route('/products/<int:product_id>/images/<int:image_id>/delete', methods=['POST'])
@admin_required
def delete_product_image(product_id, image_id):
    image = ProductImage.query.filter_by(id=image_id, product_id=product_id).first_or_404()
    if image.cloudinary_public_id:
        try:
            cloudinary.uploader.destroy(image.cloudinary_public_id)
        except Exception:
            pass
    db.session.delete(image)
    db.session.commit()
    return jsonify({'success': True})


def _handle_image_uploads(product, files, variant=None):
    """Upload Cloudinary images and attach them to a product (and optionally a variant)."""
    scope_images = variant.images if variant is not None else [
        i for i in product.images if i.variant_id is None
    ]
    is_first = not scope_images
    for f in files:
        if f and f.filename:
            try:
                result = cloudinary.uploader.upload(
                    f,
                    folder='smart99c/products',
                    transformation=[{'width': 800, 'height': 800, 'crop': 'limit', 'quality': 'auto'}]
                )
                img = ProductImage(
                    product_id=product.id,
                    variant_id=variant.id if variant else None,
                    image_url=result['secure_url'],
                    cloudinary_public_id=result['public_id'],
                    is_primary=is_first if variant is None else False,
                    sort_order=len(product.images),
                )
                db.session.add(img)
                is_first = False
            except Exception as e:
                flash(f'Image upload failed: {str(e)}', 'warning')


# ─── Product Variants ────────────────────────────────────────────────────────

def _parse_decimal(value):
    if value in (None, '', 'None'):
        return None
    try:
        return Decimal(str(value).strip())
    except Exception:
        return None


def _parse_int(value, default=0):
    if value in (None, '', 'None'):
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


@admin_bp.route('/products/<int:product_id>/variants', methods=['POST'])
@admin_required
def create_variant(product_id):
    product = Product.query.get_or_404(product_id)
    color_name = (request.form.get('color_name') or '').strip()
    if not color_name:
        flash('Color name is required.', 'error')
        return redirect(url_for('admin.edit_product', product_id=product_id))

    color_hex = (request.form.get('color_hex') or '').strip() or None
    sku = (request.form.get('sku') or '').strip() or None
    if sku and ProductVariant.query.filter_by(sku=sku).first():
        flash(f'SKU "{sku}" is already in use.', 'error')
        return redirect(url_for('admin.edit_product', product_id=product_id))

    variant = ProductVariant(
        product_id=product.id,
        color_name=color_name,
        color_hex=color_hex,
        sku=sku,
        price_override=_parse_decimal(request.form.get('price_override')),
        cost_price=_parse_decimal(request.form.get('cost_price')),
        stock_quantity=max(0, _parse_int(request.form.get('stock_quantity'), 0)),
        is_active=request.form.get('is_active') == 'on',
        sort_order=_parse_int(request.form.get('sort_order'), 0),
    )
    db.session.add(variant)
    db.session.flush()

    files = request.files.getlist('variant_images')
    _handle_image_uploads(product, files, variant=variant)

    db.session.commit()
    flash(f'Color "{variant.color_name}" added.', 'success')
    return redirect(url_for('admin.edit_product', product_id=product_id))


@admin_bp.route('/products/<int:product_id>/variants/<int:variant_id>', methods=['POST'])
@admin_required
def update_variant(product_id, variant_id):
    variant = ProductVariant.query.filter_by(id=variant_id, product_id=product_id).first_or_404()
    product = variant.product

    color_name = (request.form.get('color_name') or '').strip()
    if not color_name:
        flash('Color name is required.', 'error')
        return redirect(url_for('admin.edit_product', product_id=product_id))

    sku = (request.form.get('sku') or '').strip() or None
    if sku and sku != variant.sku:
        clash = ProductVariant.query.filter(ProductVariant.sku == sku,
                                            ProductVariant.id != variant.id).first()
        if clash:
            flash(f'SKU "{sku}" is already in use.', 'error')
            return redirect(url_for('admin.edit_product', product_id=product_id))

    variant.color_name = color_name
    variant.color_hex = (request.form.get('color_hex') or '').strip() or None
    variant.sku = sku
    variant.price_override = _parse_decimal(request.form.get('price_override'))
    variant.cost_price = _parse_decimal(request.form.get('cost_price'))
    variant.stock_quantity = max(0, _parse_int(request.form.get('stock_quantity'), 0))
    variant.is_active = request.form.get('is_active') == 'on'
    variant.sort_order = _parse_int(request.form.get('sort_order'), 0)

    files = request.files.getlist('variant_images')
    if files:
        _handle_image_uploads(product, files, variant=variant)

    db.session.commit()
    flash(f'Color "{variant.color_name}" updated.', 'success')
    return redirect(url_for('admin.edit_product', product_id=product_id))


# ─── Bulk Import / Export (Excel) ────────────────────────────────────────────

@admin_bp.route('/products/import', methods=['GET'])
@admin_required
def products_import():
    return render_template('admin/products_import.html')


@admin_bp.route('/products/import/template', methods=['GET'])
@admin_required
def products_import_template():
    from app.services.product_export import build_template_workbook, workbook_to_bytes
    wb = build_template_workbook()
    buf = workbook_to_bytes(wb)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='smart99c-products-template.xlsx')


@admin_bp.route('/products/export', methods=['GET'])
@admin_required
def products_export():
    from app.services.product_export import build_catalog_workbook, workbook_to_bytes
    wb = build_catalog_workbook()
    buf = workbook_to_bytes(wb)
    stamp = datetime.utcnow().strftime('%Y%m%d')
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'smart99c-products-{stamp}.xlsx')


@admin_bp.route('/products/import', methods=['POST'])
@admin_required
def products_import_upload():
    from app.services.product_import import import_workbook

    file = request.files.get('file')
    if not file or not file.filename:
        flash('Please choose a file to upload.', 'error')
        return redirect(url_for('admin.products_import'))

    if not file.filename.lower().endswith('.xlsx'):
        flash('Only .xlsx files are supported.', 'error')
        return redirect(url_for('admin.products_import'))

    result = import_workbook(file)
    return render_template('admin/products_import_result.html', result=result)


@admin_bp.route('/products/<int:product_id>/variants/<int:variant_id>/delete', methods=['POST'])
@admin_required
def delete_variant(product_id, variant_id):
    variant = ProductVariant.query.filter_by(id=variant_id, product_id=product_id).first_or_404()

    referenced = OrderItem.query.filter_by(variant_id=variant.id).first() is not None
    if referenced:
        variant.is_active = False
        db.session.commit()
        flash(f'Color "{variant.color_name}" has existing orders — deactivated instead of deleted.',
              'warning')
    else:
        for img in list(variant.images):
            if img.cloudinary_public_id:
                try:
                    cloudinary.uploader.destroy(img.cloudinary_public_id)
                except Exception:
                    pass
            db.session.delete(img)
        db.session.delete(variant)
        db.session.commit()
        flash('Color deleted.', 'success')
    return redirect(url_for('admin.edit_product', product_id=product_id))


# ─── Categories ──────────────────────────────────────────────────────────────

class CategoryForm(FlaskForm):
    name = StringField('Name', [validators.DataRequired(), validators.Length(max=100)])
    description = TextAreaField('Description', [validators.Optional()])
    is_active = BooleanField('Active', default=True)
    sort_order = IntegerField('Sort Order', [validators.Optional()], default=0)


@admin_bp.route('/categories')
@admin_required
def categories():
    cats = Category.query.order_by(Category.sort_order, Category.name).all()
    return render_template('admin/categories.html', categories=cats)


@admin_bp.route('/categories/new', methods=['GET', 'POST'])
@admin_required
def new_category():
    form = CategoryForm()
    if form.validate_on_submit():
        slug = generate_slug(form.name.data)
        cat = Category(
            name=form.name.data,
            slug=slug,
            description=form.description.data,
            is_active=form.is_active.data,
            sort_order=form.sort_order.data or 0,
        )
        if 'image' in request.files and request.files['image'].filename:
            try:
                result = cloudinary.uploader.upload(
                    request.files['image'],
                    folder='smart99c/categories',
                    transformation=[{'width': 600, 'height': 400, 'crop': 'fill', 'quality': 'auto'}]
                )
                cat.image_url = result['secure_url']
                cat.cloudinary_public_id = result['public_id']
            except Exception as e:
                flash(f'Image upload failed: {str(e)}', 'warning')
        db.session.add(cat)
        db.session.commit()
        flash(f'Category "{cat.name}" created!', 'success')
        return redirect(url_for('admin.categories'))
    return render_template('admin/category_form.html', form=form, category=None)


@admin_bp.route('/categories/<int:cat_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_category(cat_id):
    cat = Category.query.get_or_404(cat_id)
    form = CategoryForm(obj=cat)
    if form.validate_on_submit():
        cat.name = form.name.data
        cat.description = form.description.data
        cat.is_active = form.is_active.data
        cat.sort_order = form.sort_order.data or 0
        if 'image' in request.files and request.files['image'].filename:
            try:
                result = cloudinary.uploader.upload(
                    request.files['image'],
                    folder='smart99c/categories',
                    transformation=[{'width': 600, 'height': 400, 'crop': 'fill', 'quality': 'auto'}]
                )
                cat.image_url = result['secure_url']
                cat.cloudinary_public_id = result['public_id']
            except Exception as e:
                flash(f'Image upload failed: {str(e)}', 'warning')
        db.session.commit()
        flash(f'Category "{cat.name}" updated!', 'success')
        return redirect(url_for('admin.categories'))
    return render_template('admin/category_form.html', form=form, category=cat)


# ─── Orders ──────────────────────────────────────────────────────────────────

@admin_bp.route('/orders')
@admin_required
def orders():
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')
    search = request.args.get('q', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = Order.query
    if status:
        query = query.filter_by(status=status)
    if search:
        query = query.join(ShippingAddress, isouter=True).filter(
            db.or_(
                Order.order_number.ilike(f'%{search}%'),
                ShippingAddress.email.ilike(f'%{search}%'),
                ShippingAddress.last_name.ilike(f'%{search}%'),
            )
        )
    if date_from:
        query = query.filter(Order.created_at >= date_from)
    if date_to:
        query = query.filter(Order.created_at <= date_to + ' 23:59:59')

    from app.models import ShippingAddress
    pagination = query.order_by(Order.created_at.desc()).paginate(page=page, per_page=25)
    return render_template('admin/orders.html',
                           pagination=pagination,
                           orders=pagination.items,
                           status=status,
                           search=search)


@admin_bp.route('/orders/<int:order_id>')
@admin_required
def order_detail(order_id):
    order = Order.query.get_or_404(order_id)
    return render_template('admin/order_detail.html', order=order,
                           status_choices=Order.STATUS_CHOICES)


@admin_bp.route('/orders/<int:order_id>/status', methods=['POST'])
@admin_required
def update_order_status(order_id):
    order = Order.query.get_or_404(order_id)
    new_status = request.form.get('status')
    if new_status not in Order.STATUS_CHOICES:
        flash('Invalid status.', 'error')
        return redirect(url_for('admin.order_detail', order_id=order_id))

    if new_status == 'cancelled' and order.status not in ['cancelled', 'refunded']:
        for item in order.items:
            if item.product and item.product.track_inventory:
                if item.variant is not None:
                    item.variant.stock_quantity += item.quantity
                else:
                    item.product.stock_quantity += item.quantity

    order.status = new_status
    db.session.commit()
    flash(f'Order {order.order_number} status updated to {new_status}.', 'success')
    return redirect(url_for('admin.order_detail', order_id=order_id))


# ─── Users ───────────────────────────────────────────────────────────────────

class UserCreateForm(FlaskForm):
    first_name = StringField('First Name', [validators.DataRequired(), validators.Length(max=100)])
    last_name = StringField('Last Name', [validators.DataRequired(), validators.Length(max=100)])
    email = StringField('Email', [validators.DataRequired(), validators.Email(), validators.Length(max=255)])
    password = PasswordField('Password', [validators.DataRequired(), validators.Length(min=8)])
    is_admin = BooleanField('Admin', default=False)
    is_active = BooleanField('Active', default=True)
    force_password_change = BooleanField('Require Password Change on Next Login', default=True)


class UserEditForm(FlaskForm):
    first_name = StringField('First Name', [validators.DataRequired(), validators.Length(max=100)])
    last_name = StringField('Last Name', [validators.DataRequired(), validators.Length(max=100)])
    email = StringField('Email', [validators.DataRequired(), validators.Email(), validators.Length(max=255)])
    new_password = PasswordField('New Password', [validators.Optional(), validators.Length(min=8)])
    is_admin = BooleanField('Admin', default=False)
    is_active = BooleanField('Active', default=True)
    force_password_change = BooleanField('Require Password Change on Next Login', default=False)


@admin_bp.route('/users')
@admin_required
def users():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '')
    query = User.query
    if search:
        query = query.filter(
            db.or_(
                User.email.ilike(f'%{search}%'),
                User.first_name.ilike(f'%{search}%'),
                User.last_name.ilike(f'%{search}%'),
            )
        )
    pagination = query.order_by(User.created_at.desc()).paginate(page=page, per_page=25)
    return render_template('admin/users.html',
                           pagination=pagination,
                           users=pagination.items,
                           search=search)


@admin_bp.route('/users/new', methods=['GET', 'POST'])
@admin_required
def new_user():
    form = UserCreateForm()
    if form.validate_on_submit():
        email = form.email.data.lower().strip()
        if User.query.filter_by(email=email).first():
            flash('An account with that email already exists.', 'error')
            return render_template('admin/user_form.html', form=form, user=None)

        user = User(
            email=email,
            first_name=form.first_name.data.strip(),
            last_name=form.last_name.data.strip(),
            is_admin=form.is_admin.data,
            is_active=form.is_active.data,
            force_password_change=form.force_password_change.data,
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash(f'User "{user.full_name}" created!', 'success')
        return redirect(url_for('admin.user_detail', user_id=user.id))

    return render_template('admin/user_form.html', form=form, user=None)


@admin_bp.route('/users/<int:user_id>')
@admin_required
def user_detail(user_id):
    user = User.query.get_or_404(user_id)
    orders = Order.query.filter_by(user_id=user_id).order_by(Order.created_at.desc()).all()
    return render_template('admin/user_detail.html', user=user, orders=orders)


@admin_bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    form = UserEditForm(obj=user)

    if form.validate_on_submit():
        new_email = form.email.data.lower().strip()
        if new_email != user.email:
            existing = User.query.filter_by(email=new_email).first()
            if existing and existing.id != user.id:
                flash('An account with that email already exists.', 'error')
                return render_template('admin/user_form.html', form=form, user=user)
            user.email = new_email

        user.first_name = form.first_name.data.strip()
        user.last_name = form.last_name.data.strip()

        if user.id != current_user.id:
            user.is_admin = form.is_admin.data
            user.is_active = form.is_active.data

        user.force_password_change = form.force_password_change.data

        if form.new_password.data:
            user.set_password(form.new_password.data)

        db.session.commit()
        flash(f'User "{user.full_name}" updated!', 'success')
        return redirect(url_for('admin.user_detail', user_id=user.id))

    return render_template('admin/user_form.html', form=form, user=user)


@admin_bp.route('/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('admin.user_detail', user_id=user.id))

    has_orders = Order.query.filter_by(user_id=user.id).first() is not None
    if has_orders:
        user.is_active = False
        db.session.commit()
        flash(f'User "{user.full_name}" has existing orders and was deactivated instead of deleted.', 'warning')
        return redirect(url_for('admin.users'))

    name = user.full_name
    db.session.delete(user)
    db.session.commit()
    flash(f'User "{name}" deleted.', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/users/<int:user_id>/toggle-admin', methods=['POST'])
@admin_required
def toggle_admin(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        return jsonify({'success': False, 'message': 'Cannot change your own admin status'}), 400
    user.is_admin = not user.is_admin
    db.session.commit()
    return jsonify({'success': True, 'is_admin': user.is_admin})


@admin_bp.route('/users/<int:user_id>/toggle-active', methods=['POST'])
@admin_required
def toggle_user_active(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        return jsonify({'success': False, 'message': 'Cannot deactivate yourself'}), 400
    user.is_active = not user.is_active
    db.session.commit()
    return jsonify({'success': True, 'is_active': user.is_active})


# ─── Shipping ─────────────────────────────────────────────────────────────────

class ShippingRateForm(FlaskForm):
    name = StringField('Name', [validators.DataRequired(), validators.Length(max=100)])
    price = DecimalField('Price', [validators.DataRequired(), validators.NumberRange(min=0)], places=2)
    min_order_amount = DecimalField('Free Shipping Threshold', [validators.Optional(), validators.NumberRange(min=0)], places=2, default=0)
    estimated_days = StringField('Estimated Days', [validators.Optional(), validators.Length(max=50)])
    is_active = BooleanField('Active', default=True)
    sort_order = IntegerField('Sort Order', [validators.Optional()], default=0)


@admin_bp.route('/shipping')
@admin_required
def shipping():
    rates = ShippingRate.query.order_by(ShippingRate.sort_order, ShippingRate.price).all()
    form = ShippingRateForm()
    return render_template('admin/shipping.html', rates=rates, form=form)


@admin_bp.route('/shipping/new', methods=['POST'])
@admin_required
def new_shipping_rate():
    form = ShippingRateForm()
    if form.validate_on_submit():
        rate = ShippingRate(
            name=form.name.data,
            price=form.price.data,
            min_order_amount=form.min_order_amount.data or 0,
            estimated_days=form.estimated_days.data,
            is_active=form.is_active.data,
            sort_order=form.sort_order.data or 0,
        )
        db.session.add(rate)
        db.session.commit()
        flash(f'Shipping rate "{rate.name}" added!', 'success')
    return redirect(url_for('admin.shipping'))


@admin_bp.route('/shipping/<int:rate_id>/delete', methods=['POST'])
@admin_required
def delete_shipping_rate(rate_id):
    rate = ShippingRate.query.get_or_404(rate_id)
    db.session.delete(rate)
    db.session.commit()
    flash('Shipping rate deleted.', 'success')
    return redirect(url_for('admin.shipping'))


# ─── Settings ────────────────────────────────────────────────────────────────

@admin_bp.route('/settings', methods=['GET', 'POST'])
@admin_required
def settings():
    if request.method == 'POST':
        keys = [
            'announcement',
            'store_hours',
            'tax_rate',
            'free_shipping_threshold',
            'facebook_url',
            'instagram_url',
            'tiktok_url',
            'facebook_active',
            'instagram_active',
            'tiktok_active',
        ]
        for key in keys:
            value = '1' if key.endswith('_active') and request.form.get(key) == 'on' else request.form.get(key, '')
            setting = SiteSettings.query.filter_by(key=key).first()
            if setting:
                setting.value = value
            else:
                db.session.add(SiteSettings(key=key, value=value))
        db.session.commit()
        flash('Settings saved!', 'success')
        return redirect(url_for('admin.settings'))

    settings_dict = {s.key: s.value for s in SiteSettings.query.all()}
    return render_template('admin/settings.html', settings=settings_dict)


# ─── Email / SMTP Settings ───────────────────────────────────────────────────

@admin_bp.route('/email-settings', methods=['GET', 'POST'])
@admin_required
def email_settings():
    from app.services.mailer import send_email, is_mail_configured

    if request.method == 'POST':
        action = request.form.get('action', 'save')

        mail_server = (request.form.get('mail_server') or '').strip() or 'smtp.gmail.com'
        mail_port = (request.form.get('mail_port') or '').strip() or '587'
        mail_use_tls = '1' if request.form.get('mail_use_tls') == 'on' else '0'
        mail_username = (request.form.get('mail_username') or '').strip()
        mail_from_name = (request.form.get('mail_from_name') or '').strip()
        mail_from_email = (request.form.get('mail_from_email') or '').strip() or mail_username
        new_password = request.form.get('mail_password') or ''

        set_site_setting('mail_server', mail_server)
        set_site_setting('mail_port', mail_port)
        set_site_setting('mail_use_tls', mail_use_tls)
        set_site_setting('mail_username', mail_username)
        set_site_setting('mail_from_name', mail_from_name)
        set_site_setting('mail_from_email', mail_from_email)
        if new_password:
            set_site_setting('mail_password_encrypted', encrypt_secret(new_password))
        db.session.commit()

        if action == 'test':
            test_to = (request.form.get('test_email') or '').strip() or mail_from_email
            if not test_to:
                flash('Enter a recipient address to send a test email.', 'error')
                return redirect(url_for('admin.email_settings'))
            try:
                send_email(
                    test_to,
                    'Smart 99¢ Plus — Test Email',
                    '<p>This is a test email from your Smart 99¢ Plus admin panel.</p>'
                    f'<p>If you received this, your SMTP settings are working.</p>'
                    f'<p style="color:#888;font-size:0.85em;">Sent to {test_to}.</p>',
                    text_body=f'Test email from Smart 99¢ Plus admin panel. Sent to {test_to}.',
                )
                flash(f'Test email sent to {test_to}.', 'success')
            except Exception as e:
                flash(f'Test email failed: {e}', 'error')
        else:
            flash('Email settings saved.', 'success')
        return redirect(url_for('admin.email_settings'))

    password_is_set = bool(get_site_setting('mail_password_encrypted'))
    settings_dict = {
        'mail_server': get_site_setting('mail_server') or current_app.config.get('MAIL_SERVER', 'smtp.gmail.com'),
        'mail_port': get_site_setting('mail_port') or str(current_app.config.get('MAIL_PORT', 587)),
        'mail_use_tls': get_site_setting('mail_use_tls', '1'),
        'mail_username': get_site_setting('mail_username') or current_app.config.get('MAIL_USERNAME', ''),
        'mail_from_name': get_site_setting('mail_from_name', ''),
        'mail_from_email': get_site_setting('mail_from_email') or current_app.config.get('MAIL_USERNAME', ''),
    }
    return render_template('admin/email_settings.html',
                           settings=settings_dict,
                           password_is_set=password_is_set,
                           is_configured=is_mail_configured())


# ─── Payment Links ───────────────────────────────────────────────────────────

class PaymentLinkForm(FlaskForm):
    description = StringField('Description', [validators.DataRequired(), validators.Length(max=255)])
    amount = DecimalField('Amount (USD)', [validators.DataRequired(), validators.NumberRange(min=0.5)], places=2)
    quantity = IntegerField('Quantity', [validators.Optional(), validators.NumberRange(min=1)], default=1)
    customer_name = StringField('Customer Name (optional)', [validators.Optional(), validators.Length(max=255)])
    customer_email = StringField('Customer Email (optional)', [validators.Optional(), validators.Email(), validators.Length(max=255)])
    product_id = SelectField('Attach Product (optional)', coerce=int, default=0)
    variant_id = SelectField('Variant / Color (optional)', coerce=int, default=0)
    deduct_stock_on_paid = BooleanField('Deduct stock when paid', default=False)
    send_email = BooleanField('Email payment link to customer on create', default=False)


def _product_choices():
    products = Product.query.filter_by(is_active=True).order_by(Product.name).all()
    return [(0, '— No product —')] + [(p.id, f'{p.name} (${p.price})') for p in products]


def _variant_choices(product_id=None):
    choices = [(0, '— No variant —')]
    if product_id:
        variants = ProductVariant.query.filter_by(product_id=product_id, is_active=True)\
            .order_by(ProductVariant.sort_order, ProductVariant.id).all()
        choices += [(v.id, v.color_name) for v in variants]
    return choices


@admin_bp.route('/payment-links')
@admin_required
def payment_links():
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')
    query = PaymentLink.query
    if status:
        query = query.filter_by(status=status)
    pagination = query.order_by(PaymentLink.created_at.desc()).paginate(page=page, per_page=25)
    return render_template('admin/payment_links.html',
                           pagination=pagination,
                           links=pagination.items,
                           status=status)


@admin_bp.route('/payment-links/new', methods=['GET', 'POST'])
@admin_required
def new_payment_link():
    import stripe

    form = PaymentLinkForm()
    form.product_id.choices = _product_choices()
    # Variant choices depend on selected product. Populate based on submitted value
    # for POST validation; otherwise leave empty.
    selected_product_id = request.form.get('product_id', type=int) or request.args.get('product_id', type=int) or 0
    form.variant_id.choices = _variant_choices(selected_product_id)

    if form.validate_on_submit():
        if not current_app.config.get('STRIPE_SECRET_KEY'):
            flash('Stripe secret key is not configured. Set STRIPE_SECRET_KEY in your environment.', 'error')
            return render_template('admin/payment_link_form.html', form=form, link=None)

        product_id = form.product_id.data or None
        variant_id = form.variant_id.data or None
        if variant_id:
            variant = ProductVariant.query.get(variant_id)
            if not variant or variant.product_id != product_id:
                variant_id = None
        quantity = max(1, form.quantity.data or 1)
        amount_cents = int(Decimal(str(form.amount.data)) * 100)

        stripe.api_key = current_app.config['STRIPE_SECRET_KEY']
        try:
            stripe_product = stripe.Product.create(name=form.description.data)
            stripe_price = stripe.Price.create(
                product=stripe_product.id,
                unit_amount=amount_cents,
                currency='usd',
            )
            metadata = {'description': form.description.data}
            if product_id:
                metadata['product_id'] = str(product_id)
            if variant_id:
                metadata['variant_id'] = str(variant_id)
            metadata['quantity'] = str(quantity)
            metadata['deduct_stock'] = '1' if (form.deduct_stock_on_paid.data and product_id) else '0'
            if form.customer_email.data:
                metadata['customer_email'] = form.customer_email.data.strip()

            stripe_link = stripe.PaymentLink.create(
                line_items=[{'price': stripe_price.id, 'quantity': quantity}],
                metadata=metadata,
            )
        except Exception as e:
            flash(f'Stripe error: {e}', 'error')
            return render_template('admin/payment_link_form.html', form=form, link=None)

        link = PaymentLink(
            description=form.description.data,
            amount=form.amount.data,
            quantity=quantity,
            customer_name=(form.customer_name.data or '').strip() or None,
            customer_email=(form.customer_email.data or '').strip().lower() or None,
            product_id=product_id,
            variant_id=variant_id,
            deduct_stock_on_paid=bool(form.deduct_stock_on_paid.data and product_id),
            stripe_product_id=stripe_product.id,
            stripe_price_id=stripe_price.id,
            stripe_payment_link_id=stripe_link.id,
            url=stripe_link.url,
            status='active',
            created_by_id=current_user.id,
        )
        db.session.add(link)
        db.session.commit()

        if form.send_email.data and link.customer_email:
            _send_payment_link_email(link)

        flash(f'Payment link created: {link.url}', 'success')
        return redirect(url_for('admin.payment_link_detail', link_id=link.id))

    return render_template('admin/payment_link_form.html', form=form, link=None)


@admin_bp.route('/payment-links/variants/<int:product_id>')
@admin_required
def payment_link_variants(product_id):
    variants = ProductVariant.query.filter_by(product_id=product_id, is_active=True)\
        .order_by(ProductVariant.sort_order, ProductVariant.id).all()
    return jsonify({'variants': [{'id': v.id, 'name': v.color_name,
                                  'stock': v.stock_quantity} for v in variants]})


@admin_bp.route('/payment-links/<int:link_id>')
@admin_required
def payment_link_detail(link_id):
    link = PaymentLink.query.get_or_404(link_id)
    return render_template('admin/payment_link_detail.html', link=link)


@admin_bp.route('/payment-links/<int:link_id>/resend', methods=['POST'])
@admin_required
def resend_payment_link(link_id):
    link = PaymentLink.query.get_or_404(link_id)
    override_email = (request.form.get('customer_email') or '').strip().lower()
    if override_email:
        link.customer_email = override_email
    if not link.customer_email:
        flash('Add a customer email first.', 'error')
        return redirect(url_for('admin.payment_link_detail', link_id=link.id))
    ok, err = _send_payment_link_email(link)
    if ok:
        flash(f'Payment link emailed to {link.customer_email}.', 'success')
    else:
        flash(f'Could not send email: {err}', 'error')
    return redirect(url_for('admin.payment_link_detail', link_id=link.id))


@admin_bp.route('/payment-links/<int:link_id>/void', methods=['POST'])
@admin_required
def void_payment_link(link_id):
    import stripe
    link = PaymentLink.query.get_or_404(link_id)
    if link.status == 'paid':
        flash('Cannot void a link that has already been paid.', 'error')
        return redirect(url_for('admin.payment_link_detail', link_id=link.id))
    if current_app.config.get('STRIPE_SECRET_KEY') and link.stripe_payment_link_id:
        stripe.api_key = current_app.config['STRIPE_SECRET_KEY']
        try:
            stripe.PaymentLink.modify(link.stripe_payment_link_id, active=False)
        except Exception as e:
            flash(f'Stripe deactivation warning: {e}', 'warning')
    link.status = 'void'
    db.session.commit()
    flash('Payment link voided.', 'success')
    return redirect(url_for('admin.payment_links'))


def _send_payment_link_email(link):
    """Best-effort email send. Returns (ok, error_message)."""
    from app.services.mailer import send_email
    greeting = f'Hi {link.customer_name},' if link.customer_name else 'Hi there,'
    html = f"""
        <p>{greeting}</p>
        <p>Here is your payment link from Smart 99¢ Plus:</p>
        <p><strong>{link.description}</strong><br>
        Amount: <strong>${link.amount}</strong></p>
        <p><a href="{link.url}" style="display:inline-block;padding:10px 18px;background:#E8334A;color:#fff;text-decoration:none;border-radius:6px;">Pay now</a></p>
        <p>Or copy this URL into your browser: <br><a href="{link.url}">{link.url}</a></p>
        <p style="color:#888;font-size:0.85em;">If you did not request this, please ignore this email.</p>
    """
    text = (f'{greeting}\n\nYour payment link from Smart 99¢ Plus:\n'
            f'{link.description}\nAmount: ${link.amount}\n\n{link.url}\n')
    try:
        send_email(link.customer_email, f'Payment link: {link.description}', html, text_body=text)
        link.email_sent_at = datetime.utcnow()
        db.session.commit()
        return True, None
    except Exception as e:
        return False, str(e)
